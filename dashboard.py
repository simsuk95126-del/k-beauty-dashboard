from __future__ import annotations

import base64
import io
import os
import time
import uuid
from typing import Any

import pandas as pd
import requests
import streamlit as st
from dotenv import load_dotenv
from openai import OpenAI
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


st.set_page_config(
    page_title="Global K-Beauty Compliance",
    page_icon="💄",
    layout="wide",
)


# ============================================================
# 기본 설정
# ============================================================
load_dotenv()

APP_VERSION = "v8.0.0 (Country Review Workflow)"
API_BASE_URL = os.environ.get(
    "API_BASE_URL",
    "https://k-beauty-api.onrender.com",
).rstrip("/")
COMPLIANCE_API_URL = f"{API_BASE_URL}/api/v1/compliance-report"
DATABASE_STATUS_URL = f"{API_BASE_URL}/api/v1/database-status"

OPENAI_API_KEY = os.environ.get("OPENAI_API_KEY")
MASTER_KEY = os.environ.get("MASTER_KEY")
client = OpenAI(api_key=OPENAI_API_KEY) if OPENAI_API_KEY else None

COUNTRY_OPTIONS = {
    "US": (
        "US "
        "(Federal FDA / California Prop 65 / Toxic-Free Cosmetics Act)"
    ),
    "EU": "EU (Cosmetics Regulation EC No 1223/2009)",
    "UK": "UK (United Kingdom Cosmetics Regulation / SCPN)",
    "CN": "CN (China NMPA)",
    "ASEAN": (
        "ASEAN "
        "(Vietnam, Singapore, Thailand, Malaysia, Indonesia, Philippines, "
        "Myanmar, Cambodia, Laos, Brunei)"
    ),
    "SFDA": "SFDA (Saudi Arabia Food and Drug Authority)",
    "HALAL": (
        "HALAL "
        "(Indonesia, Malaysia, UAE, Saudi Arabia, Turkey and other markets)"
    ),
    "EAC": (
        "EAC "
        "(Eurasian Economic Union: Russia, Belarus, Kazakhstan, "
        "Armenia, Kyrgyzstan)"
    ),
}

STATUS_DISPLAY = {
    "PASS": "🟢 PASS",
    "BANNED": "🔴 BANNED",
    "RESTRICTED": "🟠 RESTRICTED",
    "WARNING_REQUIRED": "🟡 WARNING REQUIRED",
    "REVIEW_REQUIRED": "🟣 MANUAL REVIEW",
    "VERIFICATION_REQUIRED": "⚪ VERIFICATION REQUIRED",
    "REGULATED": "🔵 REGULATED",
}

OVERALL_MESSAGE = {
    "PASS": (
        "🎉 No matching regulated ingredient was found in the selected "
        "market's loaded databases."
    ),
    "FAIL": (
        "🚨 One or more ingredients matched a confirmed prohibited "
        "ingredient record."
    ),
    "RESTRICTED": (
        "⚠️ One or more ingredients are subject to confirmed restrictions "
        "or use conditions."
    ),
    "WARNING_REQUIRED": (
        "⚠️ One or more ingredients may require a warning, notification, "
        "or labeling action."
    ),
    "REVIEW_REQUIRED": (
        "🟣 One or more ingredients require manual regulatory verification "
        "before a final decision."
    ),
}


# ============================================================
# 공통 유틸리티
# ============================================================
def clean_text(value: Any, default: str = "") -> str:
    if value is None:
        return default

    try:
        if pd.isna(value):
            return default
    except (TypeError, ValueError):
        pass

    text = " ".join(str(value).strip().split())
    if text.casefold() in {"", "nan", "none", "null"}:
        return default
    return text


def normalize_status(value: Any) -> str:
    status = clean_text(value).upper()

    aliases = {
        "FAIL": "BANNED",
        "MANUAL_REVIEW": "REVIEW_REQUIRED",
        "MANUAL REVIEW": "REVIEW_REQUIRED",
        "VERIFICATION": "VERIFICATION_REQUIRED",
        "WARNING": "WARNING_REQUIRED",
    }
    return aliases.get(status, status or "VERIFICATION_REQUIRED")


def status_display(value: Any) -> str:
    status = normalize_status(value)
    return STATUS_DISPLAY.get(status, f"⚪ {status}")


def safe_int(value: Any, default: int = 0) -> int:
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return default


def safe_float(value: Any, default: float = 0.0) -> float:
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def reset_results() -> None:
    st.session_state.api_result = None
    st.session_state.review_popup_pending = False
    st.session_state.review_popup_result_id = None


def initialize_session_state() -> None:
    defaults = {
        "free_uses_left": 3,
        "api_result": None,
        "current_auth_msg": None,
        "current_auth_tier": "INVALID",
        "last_entered_key": None,
        "test_notice_shown": False,
        "disclaimer_agreed": False,
        "review_popup_pending": False,
        "review_popup_result_id": None,
        "review_popup_acknowledged_id": None,
    }

    for key, value in defaults.items():
        if key not in st.session_state:
            st.session_state[key] = value


# ============================================================
# 네트워크 및 인증
# ============================================================
def check_license_status(
    key: str,
    increment: bool = False,
) -> tuple[str, str]:
    key = clean_text(key)

    if MASTER_KEY and key == MASTER_KEY:
        return "PRO", "👑 CEO Master Key Active! (Unlimited)"

    verify_url = "https://api.gumroad.com/v2/licenses/verify"

    try:
        response = requests.post(
            verify_url,
            data={
                "product_permalink": "pkoph",
                "license_key": key,
                "increment_uses_count": "false",
            },
            timeout=10,
        )
        response.raise_for_status()
        payload = response.json()

        if (
            payload.get("success")
            and not payload.get("purchase", {}).get("refunded")
        ):
            return "PRO", "🏆 PRO Bulk Access Granted! (Unlimited)"
    except requests.RequestException:
        return (
            "ERROR",
            "📡 Connection to the license server failed. Please try again.",
        )
    except ValueError:
        return (
            "ERROR",
            "📡 The license server returned an invalid response.",
        )

    try:
        response = requests.post(
            verify_url,
            data={
                "product_permalink": "lyibre",
                "license_key": key,
                "increment_uses_count": "true" if increment else "false",
            },
            timeout=10,
        )
        response.raise_for_status()
        payload = response.json()

        if (
            payload.get("success")
            and not payload.get("purchase", {}).get("refunded")
        ):
            uses = safe_int(payload.get("uses"), 0)

            if uses >= 50:
                return (
                    "EXPIRED",
                    "🚫 Monthly limit (50/50) reached. Please upgrade to PRO.",
                )

            return (
                "STANDARD",
                f"🔓 Standard Access Granted! (Remaining: {50 - uses}/50)",
            )
    except requests.RequestException:
        return (
            "ERROR",
            "📡 Connection to the license server failed. Please try again.",
        )
    except ValueError:
        return (
            "ERROR",
            "📡 The license server returned an invalid response.",
        )

    return "INVALID", "❌ Invalid or refunded license key."


@st.cache_data(show_spinner=False, ttl=300)
def fetch_database_status() -> dict[str, Any]:
    try:
        response = requests.get(DATABASE_STATUS_URL, timeout=10)
        response.raise_for_status()
        payload = response.json()
        return payload if isinstance(payload, dict) else {}
    except (requests.RequestException, ValueError):
        return {}


@st.cache_data(show_spinner=False, ttl=3600)
def extract_ingredients_from_image(
    file_bytes: bytes,
) -> tuple[list[str], str | None]:
    if client is None:
        return [], "OPENAI_API_KEY is not configured."

    base64_image = base64.b64encode(file_bytes).decode("utf-8")
    vision_prompt = (
        "Extract cosmetic ingredient names in their original order. "
        "Return only the ingredient names separated by a vertical bar (|). "
        "Never use commas as separators because ingredient names may contain "
        "commas. Example: 정제수|글리세린|1,2-헥산디올"
    )

    try:
        response = client.chat.completions.create(
            model="gpt-4o",
            messages=[
                {
                    "role": "user",
                    "content": [
                        {"type": "text", "text": vision_prompt},
                        {
                            "type": "image_url",
                            "image_url": {
                                "url": (
                                    "data:image/jpeg;base64,"
                                    f"{base64_image}"
                                )
                            },
                        },
                    ],
                }
            ],
            temperature=0.0,
        )

        content = response.choices[0].message.content or ""
        ingredients = [
            clean_text(item)
            for item in content.split("|")
            if clean_text(item)
        ]
        return ingredients, None
    except Exception as exc:
        return [], f"Image ingredient extraction failed: {exc}"


# ============================================================
# 업로드 파일 읽기
# ============================================================
def read_uploaded_table(uploaded_file: Any) -> pd.DataFrame:
    extension = uploaded_file.name.rsplit(".", 1)[-1].casefold()

    if extension in {"xlsx", "xls"}:
        return pd.read_excel(uploaded_file)

    raw_bytes = uploaded_file.getvalue()
    encodings = ("utf-8-sig", "utf-8", "cp949", "euc-kr")

    for encoding in encodings:
        try:
            return pd.read_csv(
                io.BytesIO(raw_bytes),
                encoding=encoding,
            )
        except UnicodeDecodeError:
            continue

    return pd.read_csv(
        io.BytesIO(raw_bytes),
        encoding="utf-8",
        encoding_errors="replace",
    )


def collect_ingredients_from_files(
    uploaded_files: list[Any],
) -> tuple[dict[str, set[str]], list[str]]:
    ingredient_source_map: dict[str, set[str]] = {}
    errors: list[str] = []

    for uploaded_file in uploaded_files:
        extension = uploaded_file.name.rsplit(".", 1)[-1].casefold()

        if extension in {"jpg", "jpeg", "png"}:
            st.image(
                uploaded_file,
                caption=f"Uploaded: {uploaded_file.name}",
                width=180,
            )
            st.info(f"🤖 AI scanning: {uploaded_file.name}")

            ingredients, error = extract_ingredients_from_image(
                uploaded_file.getvalue()
            )

            if error:
                errors.append(f"{uploaded_file.name}: {error}")
                continue

            for ingredient in ingredients:
                ingredient_source_map.setdefault(
                    ingredient,
                    set(),
                ).add(uploaded_file.name)

            st.markdown(f"##### Extracted from {uploaded_file.name}")
            st.dataframe(
                pd.DataFrame(
                    {
                        "No.": range(1, len(ingredients) + 1),
                        "Ingredient": ingredients,
                    }
                ),
                hide_index=True,
                use_container_width=True,
            )
            continue

        try:
            dataframe = read_uploaded_table(uploaded_file)
        except Exception as exc:
            errors.append(f"{uploaded_file.name}: {exc}")
            continue

        if dataframe.empty or dataframe.shape[1] == 0:
            errors.append(f"{uploaded_file.name}: no readable ingredient rows")
            continue

        st.markdown(f"📊 Preview: {uploaded_file.name}")
        st.dataframe(
            dataframe.head(3),
            hide_index=True,
            use_container_width=True,
        )

        ingredients = (
            dataframe.iloc[:, 0]
            .dropna()
            .map(clean_text)
            .tolist()
        )

        for ingredient in ingredients:
            if ingredient:
                ingredient_source_map.setdefault(
                    ingredient,
                    set(),
                ).add(uploaded_file.name)

    return ingredient_source_map, errors


# ============================================================
# API 응답 변환
# ============================================================
def flatten_review_matches(
    report_details: list[dict[str, Any]],
    ingredient_source_map: dict[str, set[str]],
) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []

    for detail in report_details:
        status = normalize_status(
            detail.get("compliance_status")
            or detail.get("restriction_type")
        )

        if status != "REVIEW_REQUIRED":
            continue

        original_ingredient = clean_text(
            detail.get("original_ingredient")
        )
        inci_name = clean_text(detail.get("inci_name"))
        source_files = ", ".join(
            sorted(
                ingredient_source_map.get(
                    original_ingredient,
                    set(),
                )
            )
        )

        matches = detail.get("review_matches") or []

        if not matches:
            rows.append(
                {
                    "Original Ingredient": original_ingredient,
                    "INCI Name": inci_name,
                    "CAS Number": clean_text(
                        detail.get("cas_number"),
                        "N/A",
                    ),
                    "Suggested Status": "REGULATED",
                    "Reason": clean_text(
                        detail.get("regulation_reason")
                        or detail.get("regulation_notice")
                    ),
                    "Confidence": "",
                    "Input Source File": source_files,
                    "Regulatory Source File": "",
                    "Source Sheet": "",
                    "Source Row": "",
                    "Raw Source Text": "",
                }
            )
            continue

        for match in matches:
            confidence = safe_float(match.get("confidence"), 0.0)
            rows.append(
                {
                    "Original Ingredient": original_ingredient,
                    "INCI Name": inci_name,
                    "CAS Number": clean_text(
                        match.get("cas_number")
                        or detail.get("cas_number"),
                        "N/A",
                    ),
                    "Suggested Status": clean_text(
                        match.get("suggested_status"),
                        "REGULATED",
                    ),
                    "Reason": clean_text(
                        match.get("reason")
                        or detail.get("regulation_reason")
                    ),
                    "Confidence": confidence,
                    "Input Source File": source_files,
                    "Regulatory Source File": clean_text(
                        match.get("source_file")
                    ),
                    "Source Sheet": clean_text(
                        match.get("source_sheet")
                    ),
                    "Source Row": safe_int(
                        match.get("source_row"),
                        0,
                    )
                    or "",
                    "Raw Source Text": clean_text(
                        match.get("raw_text")
                    ),
                }
            )

    return pd.DataFrame(rows)


def build_result_dataframe(
    result_data: dict[str, Any],
    ingredient_source_map: dict[str, set[str]],
) -> pd.DataFrame:
    details = result_data.get("report_details") or []
    rows: list[dict[str, Any]] = []

    for index, detail in enumerate(details, start=1):
        original_ingredient = clean_text(
            detail.get("original_ingredient")
        )
        status = normalize_status(
            detail.get("compliance_status")
            or detail.get("restriction_type")
        )

        rows.append(
            {
                "No.": index,
                "Original Ingredient": original_ingredient,
                "INCI Name": clean_text(detail.get("inci_name")),
                "CAS Number": clean_text(
                    detail.get("cas_number"),
                    "N/A",
                ),
                "Compliance Status": status_display(status),
                "Status Code": status,
                "Source File": ", ".join(
                    sorted(
                        ingredient_source_map.get(
                            original_ingredient,
                            set(),
                        )
                    )
                ),
                "Match Source": clean_text(
                    detail.get("match_source")
                ),
                "Regulation Reason": clean_text(
                    detail.get("regulation_reason")
                ),
                "Regulation Notice": clean_text(
                    detail.get("regulation_notice")
                ),
            }
        )

    return pd.DataFrame(rows)


def determine_local_overall_status(
    result_dataframe: pd.DataFrame,
) -> str:
    if result_dataframe.empty:
        return "REVIEW_REQUIRED"

    statuses = set(
        result_dataframe["Status Code"]
        .astype(str)
        .str.upper()
    )

    if "BANNED" in statuses:
        return "FAIL"
    if "RESTRICTED" in statuses:
        return "RESTRICTED"
    if "WARNING_REQUIRED" in statuses:
        return "WARNING_REQUIRED"
    if statuses.intersection(
        {
            "REVIEW_REQUIRED",
            "VERIFICATION_REQUIRED",
            "REGULATED",
        }
    ):
        return "REVIEW_REQUIRED"
    return "PASS"


# ============================================================
# Excel 생성
# ============================================================
def format_workbook(writer: pd.ExcelWriter) -> None:
    header_fill = PatternFill(
        fill_type="solid",
        fgColor="D9EAF7",
    )
    header_font = Font(bold=True)

    for worksheet in writer.book.worksheets:
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
            )

        for column_index, column_cells in enumerate(
            worksheet.columns,
            start=1,
        ):
            max_length = 0

            for cell in column_cells:
                value = "" if cell.value is None else str(cell.value)
                max_length = max(max_length, len(value))

                if cell.row > 1:
                    cell.alignment = Alignment(
                        vertical="top",
                        wrap_text=True,
                    )

            worksheet.column_dimensions[
                get_column_letter(column_index)
            ].width = min(max(max_length + 2, 12), 60)


def create_full_report_excel(
    result_dataframe: pd.DataFrame,
    review_dataframe: pd.DataFrame,
    result_data: dict[str, Any],
) -> bytes:
    output = io.BytesIO()

    summary_rows = [
        {
            "Item": "Target Market",
            "Value": clean_text(result_data.get("target_market")),
        },
        {
            "Item": "Overall Status",
            "Value": clean_text(result_data.get("compliance_status")),
        },
        {
            "Item": "Total Checked",
            "Value": safe_int(result_data.get("total_checked")),
        },
        {
            "Item": "Banned",
            "Value": safe_int(
                (result_data.get("status_counts") or {}).get("BANNED")
            ),
        },
        {
            "Item": "Restricted",
            "Value": safe_int(
                (result_data.get("status_counts") or {}).get(
                    "RESTRICTED"
                )
            ),
        },
        {
            "Item": "Warning Required",
            "Value": safe_int(
                (result_data.get("status_counts") or {}).get(
                    "WARNING_REQUIRED"
                )
            ),
        },
        {
            "Item": "Manual Review",
            "Value": safe_int(
                (result_data.get("status_counts") or {}).get(
                    "REVIEW_REQUIRED"
                )
            ),
        },
        {
            "Item": "Verification Required",
            "Value": safe_int(
                (result_data.get("status_counts") or {}).get(
                    "VERIFICATION_REQUIRED"
                )
            ),
        },
        {
            "Item": "Confirmed DB",
            "Value": clean_text(result_data.get("database_file")),
        },
        {
            "Item": "Confirmed DB Updated",
            "Value": clean_text(
                result_data.get("database_last_updated")
            ),
        },
        {
            "Item": "Review DB",
            "Value": clean_text(
                result_data.get("review_database_file")
            ),
        },
        {
            "Item": "Review DB Updated",
            "Value": clean_text(
                result_data.get("review_database_last_updated")
            ),
        },
        {
            "Item": "Generated At",
            "Value": clean_text(
                result_data.get("report_generated_at")
            ),
        },
        {
            "Item": "Disclaimer",
            "Value": clean_text(result_data.get("disclaimer")),
        },
    ]

    report_export = result_dataframe.drop(
        columns=["Status Code"],
        errors="ignore",
    )

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        report_export.to_excel(
            writer,
            index=False,
            sheet_name="Compliance_Report",
        )
        pd.DataFrame(summary_rows).to_excel(
            writer,
            index=False,
            sheet_name="Summary",
        )

        if not review_dataframe.empty:
            review_dataframe.to_excel(
                writer,
                index=False,
                sheet_name="Manual_Review",
            )

        format_workbook(writer)

    return output.getvalue()


def create_manual_review_excel(
    review_dataframe: pd.DataFrame,
    result_data: dict[str, Any],
) -> bytes:
    output = io.BytesIO()

    instructions = pd.DataFrame(
        [
            {
                "Item": "Target Market",
                "Value": clean_text(result_data.get("target_market")),
            },
            {
                "Item": "Purpose",
                "Value": (
                    "These ingredients matched country-specific pending-"
                    "review records. They are not final legal classifications."
                ),
            },
            {
                "Item": "Required Action",
                "Value": (
                    "Verify the current official regulation, concentration "
                    "limit, product category, intended use, and source text."
                ),
            },
            {
                "Item": "Review DB",
                "Value": clean_text(
                    result_data.get("review_database_file")
                ),
            },
            {
                "Item": "Review DB Updated",
                "Value": clean_text(
                    result_data.get(
                        "review_database_last_updated"
                    )
                ),
            },
            {
                "Item": "Generated At",
                "Value": clean_text(
                    result_data.get("report_generated_at")
                ),
            },
        ]
    )

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        review_dataframe.to_excel(
            writer,
            index=False,
            sheet_name="Matched_Review_Records",
        )
        instructions.to_excel(
            writer,
            index=False,
            sheet_name="Instructions",
        )
        format_workbook(writer)

    return output.getvalue()


# ============================================================
# 결과 요약 UI
# ============================================================
def render_status_summary(
    result_dataframe: pd.DataFrame,
    result_data: dict[str, Any],
) -> None:
    status_counts = result_data.get("status_counts") or {}

    columns = st.columns(5)
    columns[0].metric(
        "PASS",
        safe_int(status_counts.get("PASS")),
    )
    columns[1].metric(
        "BANNED",
        safe_int(status_counts.get("BANNED")),
    )
    columns[2].metric(
        "RESTRICTED",
        safe_int(status_counts.get("RESTRICTED")),
    )
    columns[3].metric(
        "WARNING",
        safe_int(status_counts.get("WARNING_REQUIRED")),
    )
    columns[4].metric(
        "MANUAL REVIEW",
        safe_int(
            status_counts.get("REVIEW_REQUIRED")
        )
        + safe_int(
            status_counts.get("VERIFICATION_REQUIRED")
        ),
    )

    overall_status = clean_text(
        result_data.get("compliance_status")
    ).upper()

    if not overall_status:
        overall_status = determine_local_overall_status(
            result_dataframe
        )

    message = OVERALL_MESSAGE.get(
        overall_status,
        OVERALL_MESSAGE["REVIEW_REQUIRED"],
    )

    if overall_status == "PASS":
        st.success(message)
    elif overall_status == "FAIL":
        st.error(message)
    elif overall_status in {"RESTRICTED", "WARNING_REQUIRED"}:
        st.warning(message)
    else:
        st.info(message)


# ============================================================
# 앱 실행
# ============================================================
def run_app() -> None:
    initialize_session_state()

    @st.dialog("🚧 System Notice: Test Server")
    def show_test_server_popup() -> None:
        st.warning(
            "This system is currently running in a test environment. "
            "Temporary disruptions or incomplete data may occur."
        )
        if st.button(
            "Acknowledge",
            use_container_width=True,
            key="ack_test_notice",
        ):
            st.session_state.test_notice_shown = True
            st.rerun()

    @st.dialog("🟣 Country-Specific Manual Review Required")
    def show_manual_review_popup() -> None:
        result = st.session_state.api_result or {}
        review_dataframe = result.get("review_df", pd.DataFrame())
        target = result.get("target", "")

        st.warning(
            f"{target} pending-review data matched ingredients in this "
            "inspection. These records are not final banned or restricted "
            "classifications. Official regulations and use conditions must "
            "be checked before export, sale, or customs submission."
        )

        if not review_dataframe.empty:
            popup_columns = [
                "Original Ingredient",
                "INCI Name",
                "Suggested Status",
                "Reason",
                "Confidence",
                "Regulatory Source File",
                "Source Sheet",
                "Source Row",
            ]
            available_columns = [
                column
                for column in popup_columns
                if column in review_dataframe.columns
            ]

            st.dataframe(
                review_dataframe[available_columns],
                hide_index=True,
                use_container_width=True,
            )

        manual_review_excel = result.get("manual_review_excel")

        if manual_review_excel:
            st.download_button(
                "📥 Download This Inspection's Manual Review Report",
                data=manual_review_excel,
                file_name=f"Manual_Review_{target}.xlsx",
                mime=(
                    "application/vnd.openxmlformats-officedocument."
                    "spreadsheetml.sheet"
                ),
                use_container_width=True,
                key="popup_manual_review_download",
            )

        if st.button(
            "I understand — continue to the full result",
            use_container_width=True,
            key="ack_manual_review",
        ):
            result_id = result.get("result_id")
            st.session_state.review_popup_pending = False
            st.session_state.review_popup_acknowledged_id = result_id
            st.rerun()

    if not st.session_state.test_notice_shown:
        show_test_server_popup()
        st.stop()

    # --------------------------------------------------------
    # 사이드바: 상태 및 라이선스
    # --------------------------------------------------------
    st.sidebar.error("🚧 STATUS: TEST SERVER")
    st.sidebar.success(f"🟢 SYSTEM ONLINE ({APP_VERSION})")

    database_status = fetch_database_status()
    if database_status:
        total_confirmed = sum(
            safe_int(
                item.get("confirmed_records")
                or item.get("loaded_records")
            )
            for item in database_status.values()
            if isinstance(item, dict)
        )
        total_review = sum(
            safe_int(item.get("review_rows"))
            for item in database_status.values()
            if isinstance(item, dict)
        )
        st.sidebar.caption(
            f"Confirmed DB records loaded: {total_confirmed:,}"
        )
        st.sidebar.caption(
            f"Pending-review rows loaded: {total_review:,}"
        )
    else:
        st.sidebar.caption("Database status endpoint unavailable.")

    st.sidebar.markdown("---")
    st.sidebar.subheader("🔐 Premium Access")

    entered_password = st.sidebar.text_input(
        "Enter your License Key:",
        type="password",
    )

    is_vip = False
    is_pro_or_ceo = False

    if entered_password:
        if (
            st.session_state.last_entered_key
            != entered_password
        ):
            auth_tier, auth_msg = check_license_status(
                entered_password,
                increment=False,
            )
            st.session_state.current_auth_tier = auth_tier
            st.session_state.current_auth_msg = auth_msg
            st.session_state.last_entered_key = entered_password

        auth_tier = st.session_state.current_auth_tier
        auth_msg = st.session_state.current_auth_msg

        if auth_tier == "PRO":
            st.sidebar.success(auth_msg)
            is_vip = True
            is_pro_or_ceo = True
        elif auth_tier == "STANDARD":
            st.sidebar.success(auth_msg)
            is_vip = True
        elif auth_tier == "EXPIRED":
            st.sidebar.error(auth_msg)
        elif auth_tier == "ERROR":
            st.sidebar.warning(auth_msg)
        else:
            st.sidebar.error(auth_msg)
    else:
        st.sidebar.warning(
            "🎁 Free Trial Active: "
            f"{st.session_state.free_uses_left} checks remaining."
        )

    st.sidebar.markdown("---")
    st.sidebar.subheader("💎 Choose Your Plan")

    @st.dialog("⚠️ License Key Information & Payment")
    def show_payment_popup(
        plan_name: str,
        link: str,
    ) -> None:
        st.warning(
            "After payment, the license key will be included in the "
            "Gumroad receipt email."
        )
        st.link_button(
            f"Go to {plan_name} Payment Page",
            link,
            use_container_width=True,
        )

    if st.sidebar.button(
        "💳 Subscribe Standard ($299/mo)",
        use_container_width=True,
    ):
        show_payment_popup(
            "Standard Plan",
            "https://dahee5.gumroad.com/l/lyibre",
        )
        st.stop()

    st.sidebar.caption(
        "✔️ 50 scans per month\n\n"
        "✔️ Single-file scan and Excel download"
    )

    if st.sidebar.button(
        "🚀 Subscribe PRO Bulk ($499/mo)",
        use_container_width=True,
    ):
        show_payment_popup(
            "PRO Bulk Plan",
            "https://dahee5.gumroad.com/l/pkoph",
        )
        st.stop()

    st.sidebar.caption(
        "✔️ Unlimited scans\n\n"
        "✔️ Multiple files and batch extraction"
    )

    # --------------------------------------------------------
    # 메인 안내
    # --------------------------------------------------------
    st.title("🌍 Global K-Beauty Compliance Master")
    st.markdown(
        "##### AI-assisted ingredient normalization and country-specific "
        "regulatory screening"
    )

    advantage_columns = st.columns(2)

    with advantage_columns[0]:
        st.info(
            "#### 🛡️ Confirmed DB First\n"
            "Confirmed country-specific regulatory records take priority "
            "over pending-review records."
        )
        st.success(
            "#### 🌐 Country-Specific Review\n"
            "Only pending-review records matching this inspection and the "
            "selected market are shown."
        )

    with advantage_columns[1]:
        st.warning(
            "#### 🟣 Manual Review Separation\n"
            "Unconfirmed records are never presented as a final banned or "
            "restricted decision."
        )
        st.error(
            "#### 🇰🇷 Korean-to-INCI Matching\n"
            "The master Korean-INCI database is used before AI fallback."
        )

    st.markdown("---")

    if (
        not is_vip
        and st.session_state.free_uses_left <= 0
    ):
        st.error(
            "🔒 Free trial expired. Enter a valid license key to continue."
        )
        st.stop()

    # --------------------------------------------------------
    # 법적 고지
    # --------------------------------------------------------
    if not st.session_state.disclaimer_agreed:
        st.markdown("### ⚠️ Required Legal Notice")
        st.warning(
            "This service is an informational screening tool. It is not "
            "legal advice or an official regulatory determination. "
            "Users must independently verify current regulations, "
            "concentration limits, product categories, intended use, "
            "labeling, notification, and certification requirements."
        )

        if st.checkbox(
            "I have read and agree to the legal notice."
        ):
            st.session_state.disclaimer_agreed = True
            st.rerun()

        st.stop()

    if st.button("⚖️ Review Legal Notice"):
        st.session_state.disclaimer_agreed = False
        st.rerun()

    # --------------------------------------------------------
    # 검사 작업영역
    # --------------------------------------------------------
    st.subheader("🚀 Compliance Analysis Workspace")

    selected_display_name = st.selectbox(
        "1️⃣ Select Target Market",
        options=list(COUNTRY_OPTIONS.values()),
        on_change=reset_results,
    )

    target_country = next(
        key
        for key, value in COUNTRY_OPTIONS.items()
        if value == selected_display_name
    )

    if database_status and target_country in database_status:
        target_status = database_status[target_country]
        confirmed_count = safe_int(
            target_status.get("confirmed_records")
            or target_status.get("loaded_records")
        )
        review_count = safe_int(
            target_status.get("review_rows")
        )
        st.caption(
            f"{target_country} DB loaded: "
            f"{confirmed_count:,} confirmed records / "
            f"{review_count:,} pending-review rows"
        )

    trial_active = (
        not is_vip
        and st.session_state.free_uses_left > 0
    )

    upload_types = [
        "csv",
        "xlsx",
        "xls",
        "jpg",
        "jpeg",
        "png",
    ]

    if is_pro_or_ceo or trial_active:
        uploaded_files = st.file_uploader(
            "2️⃣ Upload Ingredient Files "
            "(Single or Multiple — PRO / Free Trial)",
            type=upload_types,
            accept_multiple_files=True,
            on_change=reset_results,
        )
    else:
        uploaded_file = st.file_uploader(
            "2️⃣ Upload One Ingredient File (Standard)",
            type=upload_types,
            accept_multiple_files=False,
            on_change=reset_results,
        )
        uploaded_files = (
            [uploaded_file]
            if uploaded_file is not None
            else []
        )

    ingredient_source_map: dict[str, set[str]] = {}

    if uploaded_files:
        ingredient_source_map, upload_errors = (
            collect_ingredients_from_files(uploaded_files)
        )

        for error in upload_errors:
            st.error(error)

        unique_ingredients = list(ingredient_source_map.keys())

        if unique_ingredients:
            st.success(
                f"{len(unique_ingredients):,} unique ingredient(s) ready."
            )

        run_clicked = (
            bool(unique_ingredients)
            and st.button(
                "🚀 Run Global Compliance Check",
                use_container_width=True,
            )
        )

        if run_clicked:
            if is_vip and not is_pro_or_ceo:
                run_tier, run_msg = check_license_status(
                    entered_password,
                    increment=True,
                )

                if run_tier == "EXPIRED":
                    st.error(
                        "Monthly limit reached. Analysis cannot proceed."
                    )
                    st.stop()

                if run_tier != "STANDARD":
                    st.error(run_msg)
                    st.stop()

                st.session_state.current_auth_msg = run_msg

            status_text = st.empty()
            progress_bar = st.progress(0)

            try:
                status_text.info(
                    "Step 1/4: Structuring ingredients and source files..."
                )
                progress_bar.progress(20)

                status_text.info(
                    f"Step 2/4: Loading {target_country} confirmed and "
                    "pending-review databases..."
                )
                progress_bar.progress(45)

                status_text.info(
                    "Step 3/4: Matching normalized INCI names..."
                )
                progress_bar.progress(70)

                response = requests.post(
                    COMPLIANCE_API_URL,
                    json={
                        "ingredients": unique_ingredients,
                        "target": target_country,
                    },
                    timeout=60,
                )

                if response.status_code != 200:
                    try:
                        error_payload = response.json()
                        error_message = error_payload.get(
                            "detail",
                            error_payload,
                        )
                    except ValueError:
                        error_message = response.text

                    raise RuntimeError(
                        f"API {response.status_code}: {error_message}"
                    )

                result_data = response.json()

                if not isinstance(result_data, dict):
                    raise RuntimeError(
                        "The API returned an invalid response structure."
                    )

                status_text.info(
                    "Step 4/4: Generating reports..."
                )
                progress_bar.progress(90)

                result_dataframe = build_result_dataframe(
                    result_data,
                    ingredient_source_map,
                )
                review_dataframe = flatten_review_matches(
                    result_data.get("report_details") or [],
                    ingredient_source_map,
                )

                full_report_excel = create_full_report_excel(
                    result_dataframe,
                    review_dataframe,
                    result_data,
                )

                manual_review_excel = (
                    create_manual_review_excel(
                        review_dataframe,
                        result_data,
                    )
                    if not review_dataframe.empty
                    else None
                )

                result_id = uuid.uuid4().hex

                st.session_state.api_result = {
                    "result_id": result_id,
                    "df": result_dataframe,
                    "review_df": review_dataframe,
                    "excel_data": full_report_excel,
                    "manual_review_excel": manual_review_excel,
                    "status": clean_text(
                        result_data.get("compliance_status")
                    ),
                    "target": target_country,
                    "result_data": result_data,
                }

                if not review_dataframe.empty:
                    st.session_state.review_popup_pending = True
                    st.session_state.review_popup_result_id = result_id
                else:
                    st.session_state.review_popup_pending = False
                    st.session_state.review_popup_result_id = None

                if not is_vip:
                    st.session_state.free_uses_left -= 1

                progress_bar.progress(100)
                status_text.success("✅ Analysis complete.")
                time.sleep(0.3)
                st.rerun()

            except (requests.RequestException, RuntimeError, ValueError) as exc:
                progress_bar.empty()
                status_text.empty()
                st.error(f"Compliance analysis failed: {exc}")

    # --------------------------------------------------------
    # 검사 결과
    # --------------------------------------------------------
    if st.session_state.api_result is not None:
        result = st.session_state.api_result
        result_dataframe = result["df"]
        review_dataframe = result["review_df"]
        result_data = result["result_data"]

        st.markdown("---")
        st.subheader("📊 Compliance Result")

        render_status_summary(
            result_dataframe,
            result_data,
        )

        all_sources = sorted(
            {
                source.strip()
                for value in result_dataframe.get(
                    "Source File",
                    pd.Series(dtype=str),
                ).astype(str)
                for source in value.split(",")
                if source.strip()
            }
        )

        selected_files = st.multiselect(
            "📂 Filter Dashboard by Input Source File",
            options=all_sources,
        )

        if selected_files:
            display_dataframe = result_dataframe[
                result_dataframe["Source File"].apply(
                    lambda value: any(
                        selected in str(value)
                        for selected in selected_files
                    )
                )
            ].copy()
            display_dataframe["No."] = range(
                1,
                len(display_dataframe) + 1,
            )
        else:
            display_dataframe = result_dataframe.copy()

        visible_columns = [
            "No.",
            "Original Ingredient",
            "INCI Name",
            "CAS Number",
            "Compliance Status",
            "Source File",
            "Regulation Reason",
            "Regulation Notice",
        ]

        st.dataframe(
            display_dataframe[
                [
                    column
                    for column in visible_columns
                    if column in display_dataframe.columns
                ]
            ],
            hide_index=True,
            use_container_width=True,
        )

        download_columns = st.columns(2)

        with download_columns[0]:
            st.download_button(
                "📥 Download Full Compliance Report",
                data=result["excel_data"],
                file_name=(
                    f"Compliance_Report_{result['target']}.xlsx"
                ),
                mime=(
                    "application/vnd.openxmlformats-officedocument."
                    "spreadsheetml.sheet"
                ),
                use_container_width=True,
            )

        with download_columns[1]:
            if (
                not review_dataframe.empty
                and result.get("manual_review_excel")
            ):
                st.download_button(
                    "🟣 Download Manual Review Report",
                    data=result["manual_review_excel"],
                    file_name=(
                        f"Manual_Review_{result['target']}.xlsx"
                    ),
                    mime=(
                        "application/vnd.openxmlformats-officedocument."
                        "spreadsheetml.sheet"
                    ),
                    use_container_width=True,
                )
            else:
                st.button(
                    "🟣 No Manual Review Matches",
                    disabled=True,
                    use_container_width=True,
                )

        if not review_dataframe.empty:
            st.markdown("### 🟣 Matched Pending-Review Records")
            st.warning(
                "Only records matching ingredients in this inspection are "
                "shown below. The complete internal review database is not "
                "exposed."
            )
            st.dataframe(
                review_dataframe,
                hide_index=True,
                use_container_width=True,
            )

        with st.expander("Database and report metadata"):
            metadata = pd.DataFrame(
                [
                    {
                        "Target Market": clean_text(
                            result_data.get("target_market")
                        ),
                        "Confirmed DB": clean_text(
                            result_data.get("database_file")
                        ),
                        "Confirmed DB Updated": clean_text(
                            result_data.get(
                                "database_last_updated"
                            )
                        ),
                        "Review DB": clean_text(
                            result_data.get(
                                "review_database_file"
                            )
                        ),
                        "Review DB Updated": clean_text(
                            result_data.get(
                                "review_database_last_updated"
                            )
                        ),
                        "Report Generated": clean_text(
                            result_data.get("report_generated_at")
                        ),
                    }
                ]
            )
            st.dataframe(
                metadata,
                hide_index=True,
                use_container_width=True,
            )
            st.caption(
                clean_text(result_data.get("disclaimer"))
            )

        result_id = result.get("result_id")
        should_show_review_popup = (
            st.session_state.review_popup_pending
            and st.session_state.review_popup_result_id == result_id
            and st.session_state.review_popup_acknowledged_id
            != result_id
            and not review_dataframe.empty
        )

        if should_show_review_popup:
            show_manual_review_popup()


if __name__ == "__main__":
    run_app()
